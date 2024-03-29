import openpyxl
import configparser
from exceptions import GetFileException, WriteFileException
import re


class ExcelWriter:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini', encoding='utf-8')

    def create_file(self) -> None:
        road = self.config.get("program", "write_in_file")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(
            ['Имя Фамилия', 'Username', 'ID', 'Номер телефона', 'Последний раз в сети', 'Бот или нет', 'Админ или нет', 'Чат'])
        worksheet.auto_filter.ref = 'A1:G1'
        workbook.save(road)
        print(f'Создан файл по пути {road}')

    def get_rows(self) -> filter:
        try:
            road = self.config.get("program", "read_file")
            rows = []
            workbook = openpyxl.load_workbook(road)

            worksheet = workbook.active

            for row in worksheet.iter_rows():
                rows.append(*[i.value for i in row])
            rows = self.__reformat(rows)
        except FileNotFoundError:
            print(f'GetFileExeption: {GetFileException.__doc__}')
            raise GetFileException
        return rows

    async def writer(self, rows) -> None:
        try:
            config = configparser.ConfigParser()
            config.read('config.ini', encoding='utf-8')
            road = config.get("program", "write_in_file")
            workbook = openpyxl.load_workbook(road)

            worksheet = workbook.active
            async for row in rows:
                worksheet.append(row)
            workbook.save(road)
        except FileNotFoundError:
            print(f'WriteFileExeption: {WriteFileException.__doc__}')
            raise WriteFileException

    @staticmethod
    def __reformat(lst: list) -> filter:
        for ind, row in enumerate(lst):
            if row:
                lst[ind] = re.sub(r'[a-zA-Z]+?://t\.me/(.+)', r'\1', row)
        return filter(None, lst)

writer = ExcelWriter()
if __name__ == '__main__':
    a = ExcelWriter()
    print(list(a.get_rows()))