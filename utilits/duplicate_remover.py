import configparser
import openpyxl
from exceptions.exceptions import GetFileException

def duplicate_remover():
    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')
    try:
        rows = []
        road = config.get("program", "write_in_file")
        workbook = openpyxl.load_workbook(road)
        worksheet = workbook.active


        for row in worksheet.iter_rows(values_only=True):
            if row not in rows:
                rows.append(row)


        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active

        for row in rows:
            new_worksheet.append(row)
        new_worksheet.auto_filter.ref = 'A1:G1'
        new_workbook.save(road)



    except FileNotFoundError:
        print(f'GetFileException: {GetFileException.__doc__}')
        raise GetFileException


